import pandas as pd
import chardet
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# 检测文件编码
def detect_encoding(file_path):
    with open(file_path, 'rb') as file:
        raw_data = file.read()
        result = chardet.detect(raw_data)
        return result['encoding']

# 读取 TXT 文件
def read_txt(file_path, encoding):
    with open(file_path, 'r', encoding=encoding) as file:
        lines = file.readlines()
    return lines

# 处理数据，按空格分割
def process_data(lines):
    data = []
    for line in lines:
        # 去除换行符并按空格分割
        row = line.strip().split()
        data.append(row)
    return data

# 将数据保存到 Excel 文件
def save_to_excel(data, output_file):
    # 将数据转换为 DataFrame
    df = pd.DataFrame(data)

    # 定义一个函数，用于标注大于 70% 的值为红色
    def highlight_gt_70(val):
        try:
            # 将值转换为百分比数值（假设值为百分比字符串，如 "74%"）
            if isinstance(val, str) and val.endswith('%'):
                num = float(val.strip('%'))
                if num > 70:
                    return 'background-color: red'
        except ValueError:
            pass
        return ''

    # 应用样式并保存到 Excel 文件
    styled_df = df.style.map(highlight_gt_70)  # 使用 map 替代 applymap
    styled_df.to_excel(output_file, index=False, header=False)


def deleteless80percentrow(filename) :
    # 加载Excel文件
    wb = load_workbook(filename)  # 请将 'your_file.xlsx' 替换为实际的文件名
    sheet = wb.active

    # 遍历每一行
    rows_to_delete = []
    for row_num in range(1, sheet.max_row + 1):
        has_percentage = False
        percentage_value = 0
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            if isinstance(cell.value, str) and "%" in cell.value:
                try:
                    percentage_value = float(cell.value.rstrip("%"))
                    has_percentage = True
                except ValueError:
                    pass
        if has_percentage:
            if percentage_value < 70:
                rows_to_delete.append(row_num)
            else:
                for col_num in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        cell.font = Font(color="FF0000")  # 设置字体颜色为红色
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                fill_type="solid")  # 设置背景色为黄色，可按需修改

    # 倒序删除行，避免行号混乱
    for row_num in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_num)

    # 保存修改后的Excel文件
    wb.save(filename)

def txt_to_excel(inputfilename, outputfilename):
    # 检测文件编码
    encoding = detect_encoding(inputfilename)
    print(f"检测到的文件编码: {encoding}")

    # 读取 TXT 文件
    lines = read_txt(inputfilename, encoding)
    # 处理数据
    data = process_data(lines)
    # 保存到 Excel 文件
    save_to_excel(data, outputfilename)

    deleteless80percentrow(outputfilename);

    print(f"数据已成功保存到 {outputfilename}")

# 主函数
def main():
    print("none")

if __name__ == "__main__":
    main()